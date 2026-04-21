"""
parsers/importacao.py
=====================
Importadores para o cadastro de produtos.

Fluxos:
  - parse_xml_nfe(xml)        → extrai itens crus da NF-e (CNPJ emit, cProd etc).
  - resolver_itens_xml(itens, cadastro)
                              → separa itens com vínculo já existente dos que
                                precisam de vinculação manual (código interno).
  - parse_xlsx_cadastro(xlsx, cadastro_existente)
                              → cria/atualiza produtos no cadastro a partir de
                                planilha com coluna 'Código Interno' obrigatória.
  - extrair_nomes_classe_distintos_xlsx(xlsx)
                              → lista nomes únicos da coluna Classe (para criar
                                classes antes do parse, na UI).
  - gerar_template_xlsx()     → template .xlsx para preenchimento manual.
  - gerar_template_precos_xlsx(resultados)
                              → template .xlsx pré-preenchido para importar
                                'Preço Praticado' em lote.
  - parse_xlsx_precos(xlsx)   → lê planilha de preços e devolve
                                {codigo_interno: preco}.
"""
from __future__ import annotations
import io
from typing import Optional

from models.produto import Produto, ParametrosGlobais

# ── Tolerância a ausência de lxml ─────────────────────────────────────────────
try:
    from lxml import etree as ET
    _LXML = True
except ImportError:
    import xml.etree.ElementTree as ET
    _LXML = False

try:
    import openpyxl
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

try:
    import pandas as pd
    _PANDAS = True
except ImportError:
    _PANDAS = False


# ─── Helpers XML ──────────────────────────────────────────────────────────────

def _tag(el) -> str:
    t = el.tag
    return t.split("}")[-1] if "}" in t else t


def _find_text(root, tag: str, default: str = "0") -> str:
    for el in root.iter():
        if _tag(el) == tag:
            return (el.text or "").strip() or default
    return default


def _find_text_local(el, tag: str, default: str = "0") -> str:
    if el is None:
        return default
    for sub in el.iter():
        if _tag(sub) == tag:
            return (sub.text or "").strip() or default
    return default


def _to_float(v: str) -> float:
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return 0.0


def _iter_tag(root, tag: str):
    for el in root.iter():
        if _tag(el) == tag:
            yield el


def _normalizar_cnpj(cnpj: str) -> str:
    return "".join(c for c in (cnpj or "") if c.isdigit())


# ─── Constantes para inferência fiscal ────────────────────────────────────────

# CST ICMS (regime normal) que envolvem Substituição Tributária
CST_COM_ST         = {"10", "30", "70", "90"}  # ST retida pelo emitente
CST_ST_PAGA_ANTES  = {"60"}                     # ICMS já cobrado anteriormente por ST

# CSOSN (Simples Nacional) que envolvem ST
CSOSN_COM_ST        = {"201", "202", "203"}
CSOSN_ST_PAGA_ANTES = {"500"}

# CFOPs de compra para revenda (entrada no estabelecimento)
CFOPS_COMPRA_REVENDA = {
    "1102", "1403", "1409", "1410", "1411",     # Internos
    "2102", "2202", "2403", "2409", "2410", "2411",  # Interestaduais
}


# ─── Parser XML NF-e ──────────────────────────────────────────────────────────

def parse_xml_nfe(source) -> tuple[dict, list[str]]:
    """
    Lê uma NF-e XML e retorna:
        { "emitente": {...}, "itens": [ {...}, ... ] }
    Cada item contém:
        cod_fornecedor, descricao, ncm, qtd,
        custo_unit, ipi_unit, frete_unit, st_unit,
        cnpj_emit, nome_emit.
    """
    avisos: list[str] = []

    try:
        if isinstance(source, str):
            tree = ET.parse(source)
            root = tree.getroot()
        elif isinstance(source, bytes):
            root = ET.fromstring(source)
        else:
            content = source.read()
            if isinstance(content, str):
                content = content.encode()
            root = ET.fromstring(content)
    except Exception as e:
        return {"emitente": {}, "itens": []}, [f"Erro ao parsear XML: {e}"]

    # Emitente e Destinatário
    cnpj_emit = ""
    nome_emit = ""
    uf_emit   = ""
    uf_dest   = ""
    for el in root.iter():
        t = _tag(el)
        if t == "emit":
            cnpj_emit = _normalizar_cnpj(_find_text_local(el, "CNPJ", ""))
            nome_emit = _find_text_local(el, "xNome", "") \
                        or _find_text_local(el, "xFant", "")
            uf_emit   = _find_text_local(el, "UF", "")
        elif t == "dest":
            uf_dest   = _find_text_local(el, "UF", "")

    if not cnpj_emit:
        avisos.append("CNPJ do emitente não encontrado no XML.")

    # idDest: 1=Operação interna | 2=Interestadual | 3=Exterior
    id_dest = _find_text(root, "idDest", "1")

    # Frete total da nota e total de produtos para rateio
    frete_total = _to_float(_find_text(root, "vFrete", "0"))
    valor_total_prod = sum(
        _to_float((el.text or "0"))
        for el in root.iter()
        if _tag(el) == "vProd"
    )

    det_list = list(_iter_tag(root, "det"))
    if not det_list:
        return ({"emitente": {"cnpj": cnpj_emit, "nome": nome_emit,
                              "uf": uf_emit},
                 "destinatario": {"uf": uf_dest},
                 "id_dest": id_dest, "itens": []},
                avisos + ["Nenhum item (<det>) encontrado no XML."])

    itens: list[dict] = []
    for det in det_list:
        prod_el = next((c for c in det if _tag(c) == "prod"),    None)
        imp_el  = next((c for c in det if _tag(c) == "imposto"), None)

        if prod_el is None:
            avisos.append("Item sem tag <prod> ignorado.")
            continue

        cod_forn  = _find_text_local(prod_el, "cProd", "")
        descricao = _find_text_local(prod_el, "xProd", "Produto sem descrição")
        ncm       = _find_text_local(prod_el, "NCM",   "")
        cfop      = _find_text_local(prod_el, "CFOP",  "")
        qtd       = _to_float(_find_text_local(prod_el, "qCom",   "1"))
        v_unit    = _to_float(_find_text_local(prod_el, "vUnCom", "0"))
        v_prod    = _to_float(_find_text_local(prod_el, "vProd",  "0"))

        # ── Impostos ────────────────────────────────────────────────────────
        cst_icms = ""
        csosn    = ""
        p_icms         = 0.0
        v_bc_st        = 0.0
        p_icms_st_xml  = 0.0
        v_icms_st      = 0.0
        p_fcp          = 0.0
        v_fcp          = 0.0
        p_fcp_st       = 0.0
        v_fcp_st       = 0.0
        ipi_total      = 0.0

        if imp_el is not None:
            for sub in imp_el.iter():
                tname = _tag(sub)
                txt = (sub.text or "").strip() if sub.text else ""
                if tname == "CST" and not cst_icms:
                    cst_icms = txt
                elif tname == "CSOSN" and not csosn:
                    csosn = txt
            p_icms        = _to_float(_find_text_local(imp_el, "pICMS",    "0"))
            v_bc_st       = _to_float(_find_text_local(imp_el, "vBCST",    "0"))
            p_mva_st      = _to_float(_find_text_local(imp_el, "pMVAST",   "0"))
            p_icms_st_xml = _to_float(_find_text_local(imp_el, "pICMSST",  "0"))
            v_icms_st     = _to_float(_find_text_local(imp_el, "vICMSST",  "0"))
            p_fcp         = _to_float(_find_text_local(imp_el, "pFCP",     "0"))
            v_fcp         = _to_float(_find_text_local(imp_el, "vFCP",     "0"))
            p_fcp_st      = _to_float(_find_text_local(imp_el, "pFCPST",   "0"))
            v_fcp_st      = _to_float(_find_text_local(imp_el, "vFCPST",   "0"))
            ipi_total     = _to_float(_find_text_local(imp_el, "vIPI",     "0"))

        ipi_unit = round(ipi_total / qtd, 4) if qtd else 0.0
        st_unit  = round(v_icms_st / qtd, 4) if qtd else 0.0

        frete_unit = 0.0
        if frete_total > 0 and valor_total_prod > 0:
            prop       = v_prod / valor_total_prod
            frete_item = frete_total * prop
            frete_unit = round(frete_item / qtd, 4) if qtd else 0.0

        itens.append({
            "cnpj_emit":       cnpj_emit,
            "nome_emit":       nome_emit,
            "uf_emit":         uf_emit,
            "uf_dest":         uf_dest,
            "id_dest":         id_dest,
            "cod_fornecedor":  cod_forn.strip(),
            "descricao":       descricao,
            "ncm":             ncm,
            "cfop":            cfop,
            "qtd":             qtd,
            "custo_unit":      v_unit,
            "ipi_unit":        ipi_unit,
            "frete_unit":      frete_unit,
            "st_unit":         st_unit,
            # Campos brutos para inferência fiscal
            "cst_icms":        cst_icms,
            "csosn":           csosn,
            "p_icms":          p_icms,
            "v_bc_st":         v_bc_st,
            "p_mva_st":        p_mva_st,
            "p_icms_st_xml":   p_icms_st_xml,
            "v_icms_st":       v_icms_st,
            "p_fcp":           p_fcp,
            "v_fcp":           v_fcp,
            "p_fcp_st":        p_fcp_st,
            "v_fcp_st":        v_fcp_st,
        })

    if not itens:
        avisos.append("Nenhum produto extraído do XML.")

    return ({"emitente": {"cnpj": cnpj_emit, "nome": nome_emit, "uf": uf_emit},
             "destinatario": {"uf": uf_dest},
             "id_dest": id_dest,
             "itens": itens},
            avisos)


# ─── Inferência fiscal a partir do XML ────────────────────────────────────────

def inferir_flags_fiscais(item: dict, params) -> dict:
    """
    Analisa um item de XML e devolve sugestões fiscais (ST / FCP / DIFAL).

    Retorna um dict com:
        sugerir_st, aliq_st_calc, motivo_st
        sugerir_fcp, aliq_fcp_calc, motivo_fcp
        sugerir_difal, aliq_difal_calc, motivo_difal
    """
    cst     = item.get("cst_icms", "") or ""
    csosn   = item.get("csosn",    "") or ""
    cfop    = item.get("cfop",     "") or ""
    id_dest = str(item.get("id_dest", "1"))
    p_icms  = float(item.get("p_icms", 0.0) or 0.0)

    v_icms_st     = float(item.get("v_icms_st",     0.0) or 0.0)
    v_fcp_st      = float(item.get("v_fcp_st",      0.0) or 0.0)
    v_bc_st       = float(item.get("v_bc_st",       0.0) or 0.0)
    p_mva_st      = float(item.get("p_mva_st",      0.0) or 0.0)
    p_icms_st_xml = float(item.get("p_icms_st_xml", 0.0) or 0.0)
    p_fcp         = float(item.get("p_fcp",         0.0) or 0.0)
    v_fcp         = float(item.get("v_fcp",         0.0) or 0.0)
    p_fcp_st_xml  = float(item.get("p_fcp_st",      0.0) or 0.0)

    # ── ST ───────────────────────────────────────────────────────────────────
    # O valor armazenado em `aliq_st` do produto passa a ser a margem MVA-ST
    # (pMVAST) lida do XML. Se o XML não trouxer pMVAST (ex.: ST por pauta
    # fixa), sugerimos ST mas sem valor — o produto ficará com aliq_st=None
    # (herda do global).
    tem_st = (
        v_icms_st > 0
        or cst in (CST_COM_ST | CST_ST_PAGA_ANTES)
        or csosn in (CSOSN_COM_ST | CSOSN_ST_PAGA_ANTES)
    )
    aliq_st_calc = p_mva_st

    motivos_st = []
    if p_mva_st > 0:   motivos_st.append(f"pMVAST={p_mva_st:.2f}%")
    if v_icms_st > 0:  motivos_st.append(f"vICMSST=R${v_icms_st:.2f}")
    if p_icms_st_xml > 0: motivos_st.append(f"pICMSST={p_icms_st_xml:.2f}%")
    if cst:            motivos_st.append(f"CST={cst}")
    if csosn:          motivos_st.append(f"CSOSN={csosn}")

    # ── FCP ──────────────────────────────────────────────────────────────────
    tem_fcp = (p_fcp > 0 or v_fcp > 0 or p_fcp_st_xml > 0 or v_fcp_st > 0)
    aliq_fcp_calc = p_fcp if p_fcp > 0 else p_fcp_st_xml

    motivos_fcp = []
    if v_fcp > 0:    motivos_fcp.append(f"vFCP=R${v_fcp:.2f}")
    if v_fcp_st > 0: motivos_fcp.append(f"vFCPST=R${v_fcp_st:.2f}")

    # ── DIFAL (heurística de compra interestadual para revenda) ─────────────
    aliq_interna = float(getattr(params, "aliq_icms_interna_destino", 18.0) or 18.0)
    sugerir_difal = (
        id_dest == "2"
        and cfop in CFOPS_COMPRA_REVENDA
        and not tem_st
        and p_icms > 0
        and p_icms < aliq_interna
    )
    aliq_difal_calc = max(0.0, aliq_interna - p_icms) if sugerir_difal else 0.0

    motivo_difal = ""
    if sugerir_difal:
        motivo_difal = (
            f"Interestadual (idDest=2), CFOP={cfop}, "
            f"pICMS NF={p_icms:.1f}% vs alíq. interna do destino "
            f"{aliq_interna:.1f}%"
        )

    return {
        "sugerir_st":      bool(tem_st),
        "aliq_st_calc":    round(aliq_st_calc, 2),
        "motivo_st":       " · ".join(motivos_st),

        "sugerir_fcp":     bool(tem_fcp),
        "aliq_fcp_calc":   round(aliq_fcp_calc, 2),
        "motivo_fcp":      " · ".join(motivos_fcp),

        "sugerir_difal":   bool(sugerir_difal),
        "aliq_difal_calc": round(aliq_difal_calc, 2),
        "motivo_difal":    motivo_difal,
    }


# ─── Resolução de itens XML contra o cadastro ────────────────────────────────

def resolver_itens_xml(
    itens: list[dict],
    cadastro: dict[str, Produto],
) -> tuple[list[dict], list[dict]]:
    """
    Separa os itens do XML em:
      - mapeados: já existe vínculo (cnpj_emit, cod_fornecedor) em algum Produto.
      - pendentes: precisam de vinculação manual pelo usuário.

    Cada item recebe o campo extra "codigo_interno" quando mapeado.
    """
    mapeados:  list[dict] = []
    pendentes: list[dict] = []

    for item in itens:
        cnpj = item.get("cnpj_emit", "")
        cod  = item.get("cod_fornecedor", "")
        achado: Optional[Produto] = None
        for prod in cadastro.values():
            if prod.tem_vinculo(cnpj, cod):
                achado = prod
                break
        if achado:
            mapeados.append({**item, "codigo_interno": achado.codigo_interno})
        else:
            pendentes.append(dict(item))

    return mapeados, pendentes


def aplicar_item_no_produto(
    produto: Produto,
    item: dict,
    aceitar_st:      bool = False,
    aceitar_fcp:     bool = False,
    aceitar_difal:   bool = False,
    aceitar_icms:    bool = False,
    nao_credita_icms: bool = False,
    aplicar_custos:  bool = True,
    flags: Optional[dict] = None,
    params: Optional[ParametrosGlobais] = None,
) -> Produto:
    """
    Atualiza custos/quantidade de um Produto com os dados de um item XML.

    Quando ``aplicar_custos=False`` (usuário rejeitou as diferenças de custo
    detectadas em um item já vinculado), os campos ``qtd``, ``custo_unitario``,
    ``ipi_unitario``, ``frete_unitario`` e ``st_unitario`` do produto são
    preservados. Vínculo, NCM e sugestões fiscais continuam sendo tratados
    normalmente.

    Para cada sugestão fiscal presente em ``flags`` (ou ``item['_flags']``):
      - ``aceitar_* = True``  → aplica valor sugerido no produto.
      - ``aceitar_* = False`` → força o bloco a usar as definições globais
        (campos do produto ficam ``None``).
    Blocos SEM sugestão no XML permanecem intocados no produto.

    Também cria o vínculo (CNPJ, cod_fornecedor) se ainda não existir.

    O preço praticado NÃO é tratado aqui — ele é persistido no par
    (produto, canal) via ``utils.estado.aplicar_preco_praticado``.
    """
    if aplicar_custos:
        produto.qtd             = float(item.get("qtd",         produto.qtd))
        produto.custo_unitario  = float(item.get("custo_unit",  produto.custo_unitario))
        produto.ipi_unitario    = float(item.get("ipi_unit",    produto.ipi_unitario))
        produto.frete_unitario  = float(item.get("frete_unit",  produto.frete_unitario))
        produto.st_unitario     = float(item.get("st_unit",     produto.st_unitario))
    if not produto.ncm and item.get("ncm"):
        produto.ncm = item["ncm"]
    produto.origem = "xml"
    produto.adicionar_vinculo(
        item.get("cnpj_emit", ""),
        item.get("cod_fornecedor", ""),
        item.get("nome_emit", ""),
    )

    flags = flags if flags is not None else item.get("_flags") or {}

    # ── ST ───────────────────────────────────────────────────────────────────
    if flags.get("sugerir_st"):
        if aceitar_st:
            produto.tem_st  = True
            aliq = flags.get("aliq_st_calc", 0.0)
            produto.aliq_st = float(aliq) if aliq else None
        else:
            produto.tem_st  = None
            produto.aliq_st = None

    # ── FCP ──────────────────────────────────────────────────────────────────
    if flags.get("sugerir_fcp"):
        if aceitar_fcp:
            aliq = flags.get("aliq_fcp_calc", 0.0)
            produto.aliq_fcp = float(aliq) if aliq else None
        else:
            produto.aliq_fcp = None

    # ── DIFAL ────────────────────────────────────────────────────────────────
    if flags.get("sugerir_difal"):
        if aceitar_difal:
            produto.tem_difal  = True
            aliq = flags.get("aliq_difal_calc", 0.0)
            produto.aliq_difal = float(aliq) if aliq else None
        else:
            produto.tem_difal  = None
            produto.aliq_difal = None

    # ── ICMS (crédito de entrada) ────────────────────────────────────────────
    # Aplica somente quando o usuário confirmou (aceitar_icms=True) e o
    # regime permite crédito de ICMS (Lucro Presumido/Real). Se o valor
    # do XML coincidir com o global, zera o override para que o produto
    # herde do global.
    if params is not None and aceitar_icms:
        permitidos = ParametrosGlobais.creditos_permitidos(params.regime)
        p_icms_xml = float(item.get("p_icms", 0.0) or 0.0)
        if permitidos.get("icms") and p_icms_xml > 0:
            aliq_global = float(params.aliq_credito_icms)
            if abs(p_icms_xml - aliq_global) <= 0.01:
                produto.aliq_credito_icms = None
            else:
                produto.aliq_credito_icms = p_icms_xml

    # Se o XML não trouxe ICMS e o usuário confirmou, desativa crédito
    # de ICMS no produto (override explícito).
    if params is not None and nao_credita_icms:
        permitidos = ParametrosGlobais.creditos_permitidos(params.regime)
        if permitidos.get("icms"):
            produto.credita_icms = False

    return produto


# ─── Planilha XLSX ───────────────────────────────────────────────────────────

COLUNAS_CADASTRO = {
    "codigo_interno":          ["Código Interno", "Codigo Interno", "Código",
                                "Codigo", "SKU", "Cód. Interno"],
    "descricao":               ["Descrição", "Descricao", "Produto"],
    "classe":                  ["Classe", "Categoria", "Grupo"],
    "ncm":                     ["NCM", "ncm"],
    "qtd":                     ["Qtd", "Quantidade"],
    "custo_unitario":          ["Custo Unit.", "Custo Unitário", "Custo"],
    "ipi_unitario":            ["IPI Unit.", "IPI"],
    "frete_unitario":          ["Frete Unit.", "Frete Rateado", "Frete"],
    "st_unitario":             ["ST Unit.", "ICMS-ST", "ST"],

    "tem_difal":               ["Tem DIFAL?", "DIFAL?", "Tem DIFAL"],
    "aliq_difal":              ["Alíq. DIFAL (%)", "Alíq DIFAL", "DIFAL %"],
    "aliq_fcp":                ["FCP (%)", "Alíq. FCP (%)", "FCP %"],

    "tem_st":                  ["Tem ST?", "ST?", "Tem ST"],
    "aliq_st":                 ["Alíq. ST (%)", "ST %"],

    "tem_antecipacao":         ["Tem Antecip.?", "Antecip.?", "Tem Antecipação"],
    "aliq_antecipacao":        ["Alíq. Antecip. (%)", "Antecip. %"],

    "credita_icms":            ["Créd. ICMS?", "Credita ICMS"],
    "aliq_credito_icms":       ["Alíq. Créd. ICMS (%)", "Créd. ICMS %"],
    "credita_pis_cofins":      ["Créd. PIS/COFINS?", "Credita PIS/COFINS"],
    "aliq_credito_pis_cofins": ["Alíq. Créd. PIS/COFINS (%)", "Créd. PIS/COFINS %"],

    "aliq_icms_interna":       ["Alíq. Interna (%)", "Alíq Interna", "ICMS Interna %"],
    "margem_desejada":         ["Margem (%)", "Margem Desejada (%)", "Margem"],

    "cnpj_fornecedor":         ["CNPJ Fornecedor", "CNPJ"],
    "cod_fornecedor":          ["Cód. Fornecedor", "Cod Fornecedor",
                                "Código Fornecedor"],
    "nome_fornecedor":         ["Fornecedor", "Nome Fornecedor"],
    "observacoes":             ["Obs.", "Observações", "Observacoes"],
}


def _normalizar_header(h: str) -> str:
    """Normaliza um nome de coluna para comparação tolerante.

    Remove asteriscos (marcadores de obrigatoriedade do template),
    espaços extras e padroniza para minúsculas.
    """
    return str(h or "").strip().lstrip("*").strip().lower()


def _mapear_colunas(headers: list[str]) -> dict[str, Optional[str]]:
    headers_lower = {_normalizar_header(h): h for h in headers if str(h or "").strip()}
    mapa: dict[str, Optional[str]] = {}
    for campo, candidatos in COLUNAS_CADASTRO.items():
        encontrado = None
        for cand in candidatos:
            chave = _normalizar_header(cand)
            if chave in headers_lower:
                encontrado = headers_lower[chave]
                break
        mapa[campo] = encontrado
    return mapa


def _cell_bool(v):
    """Converte célula de planilha em Optional[bool]. Vazio → None (usa global)."""
    if v is None:
        return None
    s = str(v).strip().lower()
    if s == "":                                     return None
    if s in ("true", "1", "sim", "s", "yes", "y"):  return True
    if s in ("false", "0", "não", "nao", "n", "no"):return False
    return None


def _cell_opt_float(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return None


def _ler_dataframe_cadastro_xlsx(source) -> tuple[Optional[object], Optional[dict], Optional[str]]:
    """Lê a planilha de cadastro e devolve ``(df, mapa_colunas)`` ou ``(None, None, erro)``.

    ``mapa_colunas`` é o retorno de :func:`_mapear_colunas` (chaves internas →
    nome da coluna na planilha).
    """
    if not _PANDAS:
        return None, None, "pandas não instalado. Execute: pip install pandas openpyxl"

    try:
        src = io.BytesIO(source) if isinstance(source, bytes) else source
        df_raw = pd.read_excel(src, dtype=str, engine="openpyxl", header=None)
    except Exception as e:
        return None, None, f"Erro ao ler Excel: {e}"

    if df_raw.empty:
        return None, None, "Planilha vazia."

    candidatos_cod = {_normalizar_header(c)
                      for c in COLUNAS_CADASTRO["codigo_interno"]}
    header_row: Optional[int] = None
    for i in range(min(len(df_raw), 15)):
        linha = [_normalizar_header(v) if not pd.isna(v) else ""
                 for v in df_raw.iloc[i].tolist()]
        if any(cell in candidatos_cod for cell in linha):
            header_row = i
            break

    if header_row is None:
        primeiras = [
            ", ".join(str(v) for v in df_raw.iloc[i].tolist()
                      if not pd.isna(v))
            for i in range(min(len(df_raw), 5))
        ]
        return None, None, (
            "Coluna 'Código Interno' não encontrada nas primeiras linhas. "
            f"Conteúdo lido: {' | '.join(primeiras)}. "
            "Baixe o template para ver o formato esperado."
        )

    headers = ["" if pd.isna(h) else str(h).strip().lstrip("*").strip()
               for h in df_raw.iloc[header_row].tolist()]

    pulou_dicas = 0
    if header_row + 1 < len(df_raw):
        proxima = [str(v).strip().lower() if not pd.isna(v) else ""
                   for v in df_raw.iloc[header_row + 1].tolist()]
        marcadores_dica = ("vazio = usa global", "obrigatório", "obrigatorio",
                           "sim/não", "sim/nao", "código ncm", "codigo ncm",
                           "opcional.", "código do produto no fornecedor",
                           "nome do produto", "observações livres",
                           "observacoes livres")
        if any(any(m in cell for m in marcadores_dica) for cell in proxima):
            pulou_dicas = 1

    df = df_raw.iloc[header_row + 1 + pulou_dicas:].copy()
    df.columns = headers
    df = df.loc[:, [c for c in df.columns if c]]
    df = df.dropna(how="all")

    mapa = _mapear_colunas(df.columns.tolist())

    if not mapa.get("codigo_interno"):
        return None, None, (
            "Coluna 'Código Interno' não encontrada. "
            f"Colunas detectadas: {', '.join(df.columns)}. "
            "Baixe o template para ver o formato esperado."
        )

    return df, mapa, None


def extrair_nomes_classe_distintos_xlsx(source) -> tuple[list[str], Optional[str]]:
    """Lista nomes únicos da coluna **Classe** (primeira ocorrência preserva o texto).

    Usado antes do parse para criar classes ainda inexistentes. Se não houver
    coluna de classe, retorna lista vazia sem erro.
    """
    df, mapa, err = _ler_dataframe_cadastro_xlsx(source)
    if err:
        return [], err
    assert df is not None and mapa is not None
    col = mapa.get("classe")
    if not col or col not in df.columns:
        return [], None

    nomes: list[str] = []
    vistos: set[str] = set()
    for _, row in df.iterrows():
        v = row[col]
        if pd.isna(v):
            continue
        s = str(v).strip()
        if not s:
            continue
        key = s.lower()
        if key not in vistos:
            vistos.add(key)
            nomes.append(s)
    return nomes, None


def parse_xlsx_cadastro(
    source,
    cadastro_existente: Optional[dict[str, Produto]] = None,
    classes_por_nome: Optional[dict[str, int]] = None,
    classe_default_id: Optional[int] = None,
) -> tuple[dict[str, Produto], list[str]]:
    """
    Lê planilha e devolve um DICT de produtos por código interno, já mesclado
    com o cadastro_existente (atualizando produtos existentes pelo código).

    Quando ``classes_por_nome`` é informado (mapa ``nome → classe_id``),
    a coluna opcional ``Classe`` da planilha vincula produtos a classes já
    cadastradas (por nome, comparação sem diferenciar maiúsculas). Para linhas
    sem classe ou quando o nome não existe no mapa, usa-se
    ``classe_default_id`` (tipicamente a classe 'Geral'). A criação automática
    de classes inexistentes fica a cargo da camada de UI antes de chamar esta
    função.
    """
    if not _PANDAS:
        return {}, ["pandas não instalado. Execute: pip install pandas openpyxl"]

    avisos: list[str] = []
    cadastro: dict[str, Produto] = dict(cadastro_existente or {})

    df, mapa, err = _ler_dataframe_cadastro_xlsx(source)
    if err:
        return cadastro, [err]

    def _raw(row, campo):
        col = mapa.get(campo)
        if col and col in row.index:
            v = row[col]
            return "" if pd.isna(v) else v
        return ""

    def _str(row, campo) -> str:
        v = _raw(row, campo)
        return "" if v == "" else str(v).strip()

    novos = 0
    atualizados = 0

    cls_map_lower = {
        (k or "").strip().lower(): v for k, v in (classes_por_nome or {}).items()
    }
    classe_avisos: set[str] = set()

    for _, row in df.iterrows():
        cod = _str(row, "codigo_interno")
        if not cod:
            continue

        existente = cadastro.get(cod)
        descricao = _str(row, "descricao")
        ncm       = _str(row, "ncm")
        classe_nome_row = _str(row, "classe")

        custo  = _cell_opt_float(_raw(row, "custo_unitario"))
        ipi    = _cell_opt_float(_raw(row, "ipi_unitario"))
        frete  = _cell_opt_float(_raw(row, "frete_unitario"))
        st_val = _cell_opt_float(_raw(row, "st_unitario"))
        qtd    = _cell_opt_float(_raw(row, "qtd"))

        if existente is None:
            produto = Produto(
                codigo_interno = cod,
                descricao      = descricao or cod,
                ncm            = ncm,
                qtd            = qtd if qtd is not None else 1.0,
                custo_unitario = custo or 0.0,
                ipi_unitario   = ipi or 0.0,
                frete_unitario = frete or 0.0,
                st_unitario    = st_val or 0.0,
                origem         = "xlsx",
            )
            novos += 1
        else:
            produto = existente
            if descricao: produto.descricao = descricao
            if ncm:       produto.ncm       = ncm
            if custo  is not None: produto.custo_unitario = custo
            if ipi    is not None: produto.ipi_unitario   = ipi
            if frete  is not None: produto.frete_unitario = frete
            if st_val is not None: produto.st_unitario    = st_val
            if qtd    is not None: produto.qtd            = qtd
            atualizados += 1

        # ── Classe (opcional) ─────────────────────────────────────────────
        if classe_nome_row:
            cid = cls_map_lower.get(classe_nome_row.strip().lower())
            if cid is not None:
                produto.classe_id = int(cid)
                produto.classe_nome = classe_nome_row.strip()
            else:
                classe_avisos.add(classe_nome_row.strip())
                if produto.classe_id is None and classe_default_id is not None:
                    produto.classe_id = int(classe_default_id)
        elif produto.classe_id is None and classe_default_id is not None:
            produto.classe_id = int(classe_default_id)

        # Parâmetros fiscais individuais (None = manter)
        for campo in (
            "tem_difal", "tem_st", "tem_antecipacao",
            "credita_icms", "credita_pis_cofins",
        ):
            v = _cell_bool(_raw(row, campo))
            if v is not None:
                setattr(produto, campo, v)

        for campo in (
            "aliq_difal", "aliq_fcp", "aliq_st", "aliq_antecipacao",
            "aliq_credito_icms", "aliq_credito_pis_cofins",
            "aliq_icms_interna", "margem_desejada",
        ):
            v = _cell_opt_float(_raw(row, campo))
            if v is not None:
                setattr(produto, campo, v)

        obs = _str(row, "observacoes")
        if obs:
            produto.observacoes = obs

        # Vínculo fornecedor (opcional por linha)
        cnpj_forn = _str(row, "cnpj_fornecedor")
        cod_forn  = _str(row, "cod_fornecedor")
        nome_forn = _str(row, "nome_fornecedor")
        if cnpj_forn and cod_forn:
            produto.adicionar_vinculo(cnpj_forn, cod_forn, nome_forn)

        cadastro[produto.codigo_interno] = produto

    if novos == 0 and atualizados == 0:
        avisos.append("Nenhuma linha com Código Interno válido foi lida.")
    else:
        avisos.append(f"Planilha processada: {novos} novo(s), "
                      f"{atualizados} atualizado(s).")

    if classe_avisos:
        nomes = ", ".join(sorted(classe_avisos))
        avisos.append(
            f"Classe(s) não cadastrada(s) e ignorada(s): {nomes}. "
            "Produtos afetados ficaram na classe padrão."
        )

    return cadastro, avisos


# ─── Gerador do template XLSX ─────────────────────────────────────────────────

def gerar_template_xlsx() -> bytes:
    """Gera bytes de um .xlsx com o template de cadastro de produtos."""
    if not _OPENPYXL:
        raise ImportError("openpyxl não instalado.")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Cadastro"

    C_NAVY  = "1F3864"
    C_GRAY  = "F2F2F2"
    C_WHITE = "FFFFFF"

    headers = [
        ("Código Interno",              "Obrigatório. Chave alfanumérica única.", True),
        ("Descrição",                   "Nome do produto.",                        False),
        ("Classe",                      "Categoria organizacional (opcional). "
                                        "Nome novo cria a classe; vazio = padrão "
                                        "da importação.",                          False),
        ("NCM",                         "Código NCM (8 dígitos).",                 False),
        ("Qtd",                         "Quantidade da última compra.",            False),
        ("Custo Unit.",                 "Custo unitário s/ impostos (R$).",        False),
        ("IPI Unit.",                   "IPI unitário da NF (R$).",                False),
        ("Frete Unit.",                 "Frete rateado unitário (R$).",            False),
        ("ST Unit.",                    "ICMS-ST pago na entrada, unitário (R$).", False),

        ("Tem DIFAL?",                  "Sim/Não. Vazio = usa global.",            False),
        ("Alíq. DIFAL (%)",             "Vazio = usa global.",                     False),
        ("FCP (%)",                     "Vazio = usa global.",                     False),
        ("Tem ST?",                     "Sim/Não. Vazio = usa global.",            False),
        ("Alíq. ST (%)",                "Vazio = usa global.",                     False),
        ("Tem Antecip.?",               "Sim/Não. Vazio = usa global.",            False),
        ("Alíq. Antecip. (%)",          "Vazio = usa global.",                     False),
        ("Créd. ICMS?",                 "Sim/Não. Vazio = usa global.",            False),
        ("Alíq. Créd. ICMS (%)",        "Vazio = usa global.",                     False),
        ("Créd. PIS/COFINS?",           "Sim/Não. Vazio = usa global.",            False),
        ("Alíq. Créd. PIS/COFINS (%)",  "Vazio = usa global.",                     False),
        ("Alíq. Interna (%)",           "ICMS interna do estado. Vazio = global.", False),
        ("Margem (%)",                  "Margem individual. Vazio = global.",      False),

        ("CNPJ Fornecedor",             "Opcional. Cria vínculo p/ import XML.",   False),
        ("Cód. Fornecedor",             "Código do produto no fornecedor (cProd).",False),
        ("Fornecedor",                  "Nome do fornecedor (opcional).",          False),
        ("Obs.",                        "Observações livres.",                     False),
    ]

    n_cols = len(headers)

    title_range = f"A1:{ws.cell(row=1, column=n_cols).column_letter}1"
    ws.merge_cells(title_range)
    c = ws["A1"]
    c.value = "CADASTRO DE PRODUTOS — PRECIFICADOR E-COMMERCE"
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=C_NAVY, end_color=C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    hint_range = f"A2:{ws.cell(row=2, column=n_cols).column_letter}2"
    ws.merge_cells(hint_range)
    c = ws["A2"]
    c.value = ("Preencha a partir da linha 4. Coluna 'Código Interno' é obrigatória. "
               "Campos vazios nos parâmetros fiscais usam os Parâmetros Globais. "
               "CNPJ + Cód. Fornecedor criam vínculo para futuras importações XML.")
    c.font = Font(name="Arial", italic=True, size=9, color="595959")
    c.fill = PatternFill("solid", start_color=C_GRAY, end_color=C_GRAY)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 6

    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    default_widths = {
        "Código Interno": 16, "Descrição": 32, "Classe": 16,
        "NCM": 12, "Qtd": 8,
        "Custo Unit.": 14, "IPI Unit.": 12, "Frete Unit.": 14, "ST Unit.": 12,
        "CNPJ Fornecedor": 18, "Cód. Fornecedor": 16, "Fornecedor": 22, "Obs.": 24,
    }

    for col, (h, hint, obrig) in enumerate(headers, 1):
        width = default_widths.get(h, 14)
        ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = width

        c4 = ws.cell(row=4, column=col, value=("* " if obrig else "") + h)
        c4.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c4.fill = PatternFill(
            "solid",
            start_color="1F3864" if obrig else "2E75B6",
            end_color  ="1F3864" if obrig else "2E75B6",
        )
        c4.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c4.border = bdr

        c5 = ws.cell(row=5, column=col, value=hint)
        c5.font = Font(name="Arial", italic=True, size=8, color="595959")
        c5.fill = PatternFill("solid", start_color=C_GRAY, end_color=C_GRAY)
        c5.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c5.border = bdr

    ws.row_dimensions[4].height = 40
    ws.row_dimensions[5].height = 32

    exemplos = [
        ["SKU-0001", "Notebook Dell Inspiron 15", "Eletrônicos", "84713012", 5, 2850.00, 0.00, 12.50, 0.00,
         "Sim", 4.0, 2.0, "", "", "", "", "", "", "", "", 18.0, 18.0,
         "12345678000199", "FORN-A-001", "Distribuidora XYZ", ""],
        ["SKU-0002", "Tênis Nike Air Max",        "Vestuário",   "64041900", 20, 180.00, 0.00, 3.50, 18.00,
         "Sim", 4.0, 2.0, "Sim", 12.0, "", "", "", "", "", "", "", 20.0,
         "", "", "", "ST na NF"],
        ["SKU-0003", "Smartphone Samsung A55",    "Geral",       "85171231", 10, 1200.00, 0.00, 8.00, 0.00,
         "", "", "", "", "", "", "", "", "", "", "", "", "",
         "", "", "", ""],
    ]
    # Coluna "Descrição" (2) e "Classe" (3) alinham à esquerda; as 4 últimas
    # (Fornecedor/Obs) são 23..26 após o acréscimo da coluna Classe.
    colunas_left = {1, 2, 3, 23, 24, 25, 26}
    for i, ex in enumerate(exemplos):
        row = 6 + i
        bg = C_WHITE if i % 2 == 0 else C_GRAY
        for col, val in enumerate(ex, 1):
            c = ws.cell(row=row, column=col, value=val if val != "" else None)
            c.font = Font(name="Arial", size=10, color="000000")
            c.fill = PatternFill("solid", start_color=bg, end_color=bg)
            c.alignment = Alignment(
                horizontal="left" if col in colunas_left else "center",
                vertical="center",
            )
            c.border = bdr
        ws.row_dimensions[row].height = 18

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ─── Importação de Preços Praticados (lote) ──────────────────────────────────

COLUNAS_PRECOS = {
    "codigo_interno": [
        "Código", "Codigo", "Código Interno", "Codigo Interno",
        "SKU", "Cód. Interno",
    ],
    "novo_preco": [
        "Novo Preço Praticado (R$)", "Novo Preço Praticado",
        "Novo Preco Praticado (R$)", "Novo Preco Praticado",
        "Novo Preço", "Novo Preco",
    ],
    "preco_praticado": [
        "Preço Praticado (R$)", "Preço Praticado",
        "Preco Praticado (R$)", "Preco Praticado",
        "Preço", "Preco",
    ],
}


def _mapear_colunas_precos(headers: list[str]) -> dict[str, Optional[str]]:
    headers_lower = {h.strip().lower(): h for h in headers}
    mapa: dict[str, Optional[str]] = {}
    for campo, candidatos in COLUNAS_PRECOS.items():
        encontrado = None
        for cand in candidatos:
            if cand.strip().lower() in headers_lower:
                encontrado = headers_lower[cand.strip().lower()]
                break
        mapa[campo] = encontrado
    return mapa


def gerar_template_precos_xlsx(resultados: list, canal=None) -> bytes:
    """
    Gera bytes de um .xlsx pré-preenchido com os produtos atualmente
    precificados. O usuário só precisa editar a coluna
    ``Novo Preço Praticado (R$)`` e reimportar.

    Quando ``canal`` é fornecido, o nome do canal aparece no título do
    relatório para deixar explícito que os preços se aplicam a ele.

    Colunas:
      - Código              (chave — não editar)
      - Produto             (descrição — referência)
      - Preço Mínimo (R$)   (referência)
      - Preço Praticado Atual (R$) (referência)
      - Novo Preço Praticado (R$)  (editar aqui)
    """
    if not _OPENPYXL:
        raise ImportError("openpyxl não instalado.")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Preços"

    C_NAVY   = "1F3864"
    C_BLUE   = "2E75B6"
    C_GRAY   = "F2F2F2"
    C_WHITE  = "FFFFFF"
    C_YELLOW = "FFF2CC"

    headers = [
        ("Código",                       16, "center"),
        ("Produto",                      36, "left"),
        ("Preço Mínimo (R$)",            18, "center"),
        ("Preço Praticado Atual (R$)",   22, "center"),
        ("Novo Preço Praticado (R$)",    22, "center"),
    ]
    n_cols = len(headers)

    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    titulo = "PREÇOS PRATICADOS — PRECIFICADOR E-COMMERCE"
    if canal is not None:
        titulo += f" — CANAL: {str(getattr(canal, 'nome', '')).upper()}"
    ws.merge_cells(f"A1:{ws.cell(row=1, column=n_cols).column_letter}1")
    c = ws["A1"]
    c.value = titulo
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=C_NAVY, end_color=C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Dica
    ws.merge_cells(f"A2:{ws.cell(row=2, column=n_cols).column_letter}2")
    c = ws["A2"]
    c.value = ("Edite apenas a coluna 'Novo Preço Praticado (R$)'. "
               "O produto é identificado pela coluna 'Código'. "
               "Linhas em branco ou com preço ≤ 0 são ignoradas. "
               "Valores abaixo do Preço Mínimo são aplicados com aviso.")
    c.font = Font(name="Arial", italic=True, size=9, color="595959")
    c.fill = PatternFill("solid", start_color=C_GRAY, end_color=C_GRAY)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 6

    # Cabeçalhos (linha 4)
    for col, (h, width, _) in enumerate(headers, 1):
        letter = ws.cell(row=4, column=col).column_letter
        ws.column_dimensions[letter].width = width

        is_editavel = (h == "Novo Preço Praticado (R$)")
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill(
            "solid",
            start_color=C_NAVY if is_editavel else C_BLUE,
            end_color  =C_NAVY if is_editavel else C_BLUE,
        )
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr
    ws.row_dimensions[4].height = 36

    # Dados
    MOEDA = 'R$ #,##0.00'

    for i, r in enumerate(resultados):
        row = 5 + i
        n = i + 1
        bg = C_WHITE if n % 2 else C_GRAY
        p = r.produto

        preco_min   = float(r.preco_minimo)
        preco_atual = float(r.preco_praticado)

        vals = [
            p.codigo_interno,
            p.descricao,
            preco_min,
            preco_atual,
            preco_atual,  # Novo Preço começa igual ao atual — basta editar
        ]

        for col, (val, (h, _, align_h)) in enumerate(zip(vals, headers), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=10, color="000000")
            is_editavel = (h == "Novo Preço Praticado (R$)")
            c.fill = PatternFill(
                "solid",
                start_color=C_YELLOW if is_editavel else bg,
                end_color  =C_YELLOW if is_editavel else bg,
            )
            c.alignment = Alignment(horizontal=align_h, vertical="center")
            c.border = bdr
            if col in (3, 4, 5):
                c.number_format = MOEDA
        ws.row_dimensions[row].height = 17

    ws.freeze_panes = "A5"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def parse_xlsx_precos(source) -> tuple[dict[str, float], list[str]]:
    """
    Lê planilha de preços e devolve ``(precos_por_codigo, avisos)``.

    - ``precos_por_codigo``: ``dict[str, float]`` mapeando código interno para
      o novo preço praticado. Contém apenas linhas com código não vazio e
      preço > 0.
    - ``avisos``: lista de mensagens (coluna faltante, linhas ignoradas,
      duplicatas, etc.).

    Aceita como fonte do preço a coluna ``Novo Preço Praticado (R$)``
    (preferencial) ou ``Preço Praticado (R$)`` (fallback).
    """
    if not _PANDAS:
        return {}, ["pandas não instalado. Execute: pip install pandas openpyxl"]

    avisos: list[str] = []

    try:
        if isinstance(source, bytes):
            source = io.BytesIO(source)
        df = pd.read_excel(source, dtype=str, engine="openpyxl", header=None)
    except Exception as e:
        return {}, [f"Erro ao ler Excel: {e}"]

    if df.empty:
        return {}, ["Planilha vazia."]

    # Encontra a linha de cabeçalho: primeira que contenha alguma coluna
    # candidata de código (aceita o template com título nas linhas 1–2).
    header_row = None
    for i in range(min(len(df), 10)):
        linha = [
            "" if pd.isna(v) else str(v).strip().lower()
            for v in df.iloc[i].tolist()
        ]
        candidatos = {c.strip().lower() for c in COLUNAS_PRECOS["codigo_interno"]}
        if any(cell in candidatos for cell in linha):
            header_row = i
            break

    if header_row is None:
        return {}, [
            "Não foi possível encontrar uma linha de cabeçalho com a coluna "
            "'Código'. Baixe o template para ver o formato esperado."
        ]

    headers_raw = df.iloc[header_row].tolist()
    headers = [
        "" if pd.isna(h) else str(h).strip()
        for h in headers_raw
    ]
    dados = df.iloc[header_row + 1:].copy()
    dados.columns = headers
    dados = dados.loc[:, [c for c in dados.columns if c]]
    dados = dados.dropna(how="all")

    mapa = _mapear_colunas_precos(dados.columns.tolist())

    if not mapa.get("codigo_interno"):
        return {}, [
            "Coluna 'Código' não encontrada. "
            f"Colunas detectadas: {', '.join(dados.columns)}."
        ]

    col_novo     = mapa.get("novo_preco")
    col_fallback = mapa.get("preco_praticado")
    if not col_novo and not col_fallback:
        return {}, [
            "Nenhuma coluna de preço encontrada. Esperado "
            "'Novo Preço Praticado (R$)' ou 'Preço Praticado (R$)'."
        ]

    col_cod = mapa["codigo_interno"]

    precos: dict[str, float] = {}
    duplicados: list[str] = []
    linhas_invalidas = 0

    for _, row in dados.iterrows():
        cod_raw = row.get(col_cod, "")
        cod = "" if pd.isna(cod_raw) else str(cod_raw).strip()
        if not cod:
            continue

        valor_raw = None
        if col_novo:
            v = row.get(col_novo)
            if not (v is None or (isinstance(v, float) and pd.isna(v))
                    or str(v).strip() == ""):
                valor_raw = v
        if valor_raw is None and col_fallback:
            v = row.get(col_fallback)
            if not (v is None or (isinstance(v, float) and pd.isna(v))
                    or str(v).strip() == ""):
                valor_raw = v

        preco = _cell_opt_float(valor_raw)
        if preco is None or preco <= 0:
            linhas_invalidas += 1
            continue

        if cod in precos:
            duplicados.append(cod)
        precos[cod] = preco

    if duplicados:
        unicos = sorted(set(duplicados))
        avisos.append(
            f"Códigos duplicados na planilha (último valor prevalece): "
            f"{', '.join(unicos)}"
        )
    if linhas_invalidas:
        avisos.append(
            f"{linhas_invalidas} linha(s) ignorada(s) por preço ausente, "
            "não numérico ou ≤ 0."
        )
    if not precos:
        avisos.append("Nenhum preço válido encontrado na planilha.")

    return precos, avisos
