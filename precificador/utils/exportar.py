"""
utils/exportar.py
=================
Gera o Excel de resultado da precificação com formatação profissional.
"""
from __future__ import annotations
import io
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from models.produto import ResultadoPrecificacao, ParametrosGlobais

try:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter
    _OK = True
except ImportError:
    _OK = False


def _fl(color: str):
    return PatternFill("solid", start_color=color, end_color=color)

def _fnt(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _al(h="center", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

def _bdr(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def exportar_resultado_xlsx(
    resultados: list,          # list[ResultadoPrecificacao]
    params,                    # ParametrosGlobais
    canal=None,                # CanalVenda | None
    empresa: dict | None = None,
) -> bytes:
    """
    Gera bytes de um arquivo .xlsx com:
    - Aba Resumo Parâmetros (empresa) + resumo do canal, quando informado
    - Aba Precificação detalhada
    """
    if not _OK:
        raise ImportError("openpyxl não instalado.")

    wb = Workbook()

    # ── Aba 1: Parâmetros ─────────────────────────────────────────────────────
    ws_p = wb.active
    ws_p.title = "Parâmetros"
    ws_p.sheet_view.showGridLines = False
    ws_p.sheet_properties.tabColor = "1F3864"

    ws_p.column_dimensions["A"].width = 42
    ws_p.column_dimensions["B"].width = 22

    ws_p.merge_cells("A1:B1")
    c = ws_p["A1"]
    titulo = "PARÂMETROS UTILIZADOS NA PRECIFICAÇÃO"
    if canal is not None:
        titulo += f" — CANAL: {canal.nome.upper()}"
    c.value = titulo
    c.font = _fnt(bold=True, color="FFFFFF", size=12)
    c.fill = _fl("1F3864"); c.alignment = _al("center","center")
    ws_p.row_dimensions[1].height = 28

    resumo: dict[str, str] = {}
    resumo.update(params.resumo())
    if canal is not None:
        resumo.update(canal.resumo(params))
    for i, (lbl, val) in enumerate(resumo.items(), 2):
        ws_p.row_dimensions[i].height = 18
        c1 = ws_p.cell(row=i, column=1, value=lbl)
        c1.font = _fnt(color="000000"); c1.fill = _fl("DEEAF1" if i%2 else "FFFFFF")
        c1.alignment = _al("left", indent=1); c1.border = _bdr()
        c2 = ws_p.cell(row=i, column=2, value=val)
        c2.font = _fnt(bold=True, color="0000FF"); c2.fill = _fl("FFF2CC")
        c2.alignment = _al("center"); c2.border = _bdr()

    # ── Aba 2: Precificação ───────────────────────────────────────────────────
    ws_r = wb.create_sheet("Precificação")
    ws_r.sheet_view.showGridLines = False
    ws_r.sheet_properties.tabColor = "375623"
    ws_r.freeze_panes = "A3"

    campos = [
        ("Código",                    14, "center"),
        ("Produto",                  30, "left"),
        ("Classe",                    16, "left"),
        ("NCM",                       12, "center"),
        ("Custo Base (R$)",           14, "center"),
        ("DIFAL (R$)",                12, "center"),
        ("FCP (R$)",                  11, "center"),
        ("Antecipação (R$)",          14, "center"),
        ("(-) Crédito (R$)",          14, "center"),
        ("Custo Final (R$)",          14, "center"),
        ("Custos Op. (R$)",           13, "center"),
        ("% Impostos",                12, "center"),
        ("% ICMS embutido no DAS",    16, "center"),
        ("% Comissão",                12, "center"),
        ("% Financeiro",              12, "center"),
        ("Margem Desejada",           14, "center"),
        ("Preço Mínimo (R$)",         16, "center"),
        ("Preço Praticado (R$)",      16, "center"),
        ("Lucro Líq. Unit. (R$)",     16, "center"),
        ("Markup s/ Custo (%)",       15, "center"),
        ("Margem Líq. Real (%)",      15, "center"),
        ("Status",                    14, "center"),
    ]

    # Título
    ws_r.merge_cells(f"A1:{get_column_letter(len(campos))}1")
    c = ws_r["A1"]
    partes = ["PRECIFICAÇÃO"]
    if empresa is not None:
        nome_emp = str(empresa.get("nome") or "").strip()
        if nome_emp:
            partes.append(nome_emp.upper())
    if canal is not None:
        partes.append(str(canal.nome).upper())
    c.value = " ".join(partes) + " — RESULTADO DETALHADO"
    c.font = _fnt(bold=True, color="FFFFFF", size=12)
    c.fill = _fl("375623"); c.alignment = _al("center","center")
    ws_r.row_dimensions[1].height = 28

    # Cabeçalhos
    ws_r.row_dimensions[2].height = 42
    for col, (h, w, _) in enumerate(campos, 1):
        ws_r.column_dimensions[get_column_letter(col)].width = w
        c = ws_r.cell(row=2, column=col, value=h)
        c.font = _fnt(bold=True, color="FFFFFF", size=9)
        c.fill = _fl("1F3864"); c.alignment = _al("center","center",wrap=True)
        c.border = _bdr()

    # Dados
    MOEDA   = 'R$ #,##0.00'
    MOEDA4  = 'R$ #,##0.0000'
    PERC    = '0.00%'
    PERC1   = '0.0%'

    fmts = [
        None, None, None, None,
        MOEDA4, MOEDA4, MOEDA4, MOEDA4, MOEDA4, MOEDA4, MOEDA,
        PERC, PERC, PERC, PERC, PERC,
        MOEDA, MOEDA, MOEDA,
        PERC1, PERC1, None,
    ]

    for i, res in enumerate(resultados):
        row = i + 3
        n   = i + 1
        bg  = "FFFFFF" if n % 2 else "F2F2F2"
        ws_r.row_dimensions[row].height = 17

        d = res.to_dict()
        vals = [
            d["Código"],
            d["Produto"],
            d.get("Classe", ""),
            d["NCM"],
            d["Custo Base (R$)"],
            d["DIFAL (R$)"],
            d["FCP (R$)"],
            d["Antecipação (R$)"],
            d["(-) Crédito ICMS (R$)"] + d["(-) Crédito PIS/COF (R$)"],
            d["Custo Final (R$)"],
            d["Custos Op. (R$)"],
            d["% Impostos s/ Venda"],
            float(params.aliq_icms_proprio) / 100,
            d["% Comissão+Gateway"],
            d["% Financeiro"],
            d["Margem Desejada"],
            d["Preço Mínimo (R$)"],
            d["Preço Praticado (R$)"],
            d["Lucro Líq. Unit. (R$)"],
            d["Markup s/ Custo (%)"] / 100,
            d["Margem Líq. Real (%)"] / 100,
            d["Status"],
        ]

        for col, (val, (_, _, align_h), fmt) in enumerate(zip(vals, campos, fmts), 1):
            c = ws_r.cell(row=row, column=col, value=val)
            # Cor especial para colunas de resultado (a partir de "Preço Mínimo").
            if col >= 17:
                cell_bg = "E2EFDA" if n%2 else "C6EFCE"
            else:
                cell_bg = bg
            # "Margem Líq. Real" está na coluna 21.
            if col == 21 and isinstance(val, float) and val < 0:
                c.font = _fnt(bold=True, color="9C0006")
                cell_bg = "FFC7CE"
            elif col == 21 and isinstance(val, float) and val >= 0:
                c.font = _fnt(bold=True, color="276221")
            else:
                c.font = _fnt(color="000000" if col not in [1] else "000000")
            c.fill = _fl(cell_bg)
            c.alignment = _al(align_h, "center")
            c.border = _bdr()
            if fmt:
                c.number_format = fmt

    # ── Aba 3: Resumo por Classe ──────────────────────────────────────────────
    ws_c = wb.create_sheet("Resumo por Classe")
    ws_c.sheet_view.showGridLines = False
    ws_c.sheet_properties.tabColor = "7030A0"
    ws_c.freeze_panes = "A3"

    campos_cls = [
        ("Classe",            24, "left"),
        ("Qtd. Produtos",     14, "center"),
        ("Custo Médio (R$)",  16, "center"),
        ("Preço Médio (R$)",  16, "center"),
        ("Margem Real Média", 18, "center"),
        ("Markup Médio",      16, "center"),
        ("✅ OK",             10, "center"),
        ("⚠️ Abaixo",        12, "center"),
        ("🔴 Prejuízo",      12, "center"),
        ("% OK",              10, "center"),
        ("% Prejuízo",        12, "center"),
    ]
    fmts_cls = [
        None, None,
        MOEDA, MOEDA,
        PERC1, PERC1,
        None, None, None,
        PERC1, PERC1,
    ]

    ws_c.merge_cells(f"A1:{get_column_letter(len(campos_cls))}1")
    c = ws_c["A1"]
    partes = ["RESUMO POR CLASSE"]
    if empresa is not None:
        nome_emp = str(empresa.get("nome") or "").strip()
        if nome_emp:
            partes.append(nome_emp.upper())
    if canal is not None:
        partes.append(str(canal.nome).upper())
    c.value = " — ".join(partes)
    c.font = _fnt(bold=True, color="FFFFFF", size=12)
    c.fill = _fl("7030A0"); c.alignment = _al("center", "center")
    ws_c.row_dimensions[1].height = 28

    ws_c.row_dimensions[2].height = 38
    for col, (h, w, _) in enumerate(campos_cls, 1):
        ws_c.column_dimensions[get_column_letter(col)].width = w
        c = ws_c.cell(row=2, column=col, value=h)
        c.font = _fnt(bold=True, color="FFFFFF", size=9)
        c.fill = _fl("7030A0")
        c.alignment = _al("center", "center", wrap=True)
        c.border = _bdr()

    # Agrega por classe
    agregados: dict[str, dict] = {}
    for res in resultados:
        cls = (getattr(res.produto, "classe_nome", "") or "(sem classe)")
        a = agregados.setdefault(cls, {
            "n": 0, "custo": 0.0, "preco": 0.0,
            "margem": 0.0, "markup": 0.0,
            "ok": 0, "warn": 0, "bad": 0,
        })
        a["n"]      += 1
        a["custo"]  += float(res.custo_final)
        a["preco"]  += float(res.preco_praticado)
        a["margem"] += float(res.margem_liquida_real)
        a["markup"] += float(res.markup_sobre_custo)
        if "✅" in res.status: a["ok"]   += 1
        if "⚠️" in res.status: a["warn"] += 1
        if "🔴" in res.status: a["bad"]  += 1

    for i, cls in enumerate(sorted(agregados.keys())):
        a = agregados[cls]
        n = a["n"] or 1
        row = 3 + i
        ws_c.row_dimensions[row].height = 18
        bg = "FFFFFF" if (i + 1) % 2 else "F2F2F2"
        vals = [
            cls,
            a["n"],
            a["custo"]  / n,
            a["preco"]  / n,
            a["margem"] / n,
            a["markup"] / n,
            a["ok"],
            a["warn"],
            a["bad"],
            (a["ok"]  / a["n"]) if a["n"] else 0.0,
            (a["bad"] / a["n"]) if a["n"] else 0.0,
        ]
        for col, (val, (_, _, align_h), fmt) in enumerate(
            zip(vals, campos_cls, fmts_cls), 1
        ):
            c = ws_c.cell(row=row, column=col, value=val)
            c.font = _fnt(color="000000", bold=(col == 1))
            c.fill = _fl(bg)
            c.alignment = _al(align_h, "center")
            c.border = _bdr()
            if fmt:
                c.number_format = fmt

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
